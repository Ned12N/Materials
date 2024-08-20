import openpyxl
import pandas as pd
from openpyxl.styles import Border, Side, Font ,Alignment


def sort_and_save_sheets(input_file: str, output_file: str, specified_columns: list[str] = [] ,specified_sheetname: list[str] = [], ascending : bool = True) -> None:
# def sort_and_save_sheets(input_file, output_file , specified_columns=None ):
    """
    Reads an Excel file, sorts each sheet by all column headers,
    and saves the sorted sheets to a new Excel file.

    Parameters:
    input_file (str): Path to the input Excel file.
    output_file (str): Path to the output Excel file.
    """
    sorted_sheets = {}

    # Read the Excel file
    xl = pd.ExcelFile(input_file)

    # Loop through each sheet 
    for sheet_name in xl.sheet_names:
        
        # Check if the sheet needs to be sorted based on specified_sheetname
        if not specified_sheetname or sheet_name in specified_sheetname:

            # Read the sheet data into a DataFrame
            df = xl.parse(sheet_name)

            # If specified_columns is provided, sort by those columns; otherwise, sort by all columns
            if specified_columns:
                columns_to_sort_by = [col for col in specified_columns if col in df.columns]
            else:
                columns_to_sort_by = list(df.columns)

            # Sort the DataFrame by the specified columns and order and store it in Dictionary
            sorted_df = df.sort_values(by=columns_to_sort_by , ascending=ascending)
            sorted_sheets[sheet_name] = sorted_df
        else:
            # If the sheet doesn't need sorting, just store it as is
            sorted_sheets[sheet_name] = xl.parse(sheet_name)

    # Save all sorted sheets back to the output Excel file
    with pd.ExcelWriter(output_file) as writer:
        for sheet, sorted_data in sorted_sheets.items():
            sorted_data.to_excel(writer, sheet_name=sheet, index=False)

    print(f"Sorted data saved to {output_file}")

def autofit_columns_and_rows(excel_file):
    """
    Applies AutoFit to all columns and rows in all sheets of the given Excel file
    and centers the content within each cell.

    Parameters:
    excel_file (str): Path to the Excel file to be adjusted.
    """
    # Load the workbook and iterate through each sheet
    wb = openpyxl.load_workbook(excel_file)
    for sheet in wb.worksheets:
        # AutoFit for columns
        for column_cells in sheet.columns:
            max_length = 0
            column = column_cells[0].column_letter  # Get the column name
            for cell in column_cells:
                try:
                    # Calculate the maximum length of the column content
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                    
                    # Center the content within the cell
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
        
        # AutoFit for rows (optional)
        for row_cells in sheet.iter_rows():
            max_height = 15  # Set a default row height (if needed)
            for cell in row_cells:
                # Adjust row height based on the font size or other factors if required
                # Currently, this example sets the alignment only.
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # You can adjust row height if necessary:
            # sheet.row_dimensions[cell.row].height = adjusted_height

    # Save the workbook with the AutoFit and alignment applied
    wb.save(excel_file)
    print(f"AutoFit and centering applied to all columns and rows in {excel_file}")


def add_borders_and_bold_headers_to_sheets(excel_file: str, output_file: str) -> None:
    """
    Adds borders to all cells and makes headers (first row) bold in all sheets of the given Excel file.
    Saves the result to a new Excel file.

    Parameters:
    excel_file (str): Path to the input Excel file.
    output_file (str): Path to the output Excel file.
    """
    # Load the workbook and iterate through each sheet
    wb = openpyxl.load_workbook(excel_file)
    
    # Define the border style
    thin_border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )

    # Define the bold font style
    bold_font = Font(bold=True)

    for sheet in wb.worksheets:
        # Apply borders to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Make the header (first row) bold
        for cell in sheet[1]:  # sheet[1] refers to the first row
            cell.font = bold_font

    # Save the workbook with borders and bold headers applied
    wb.save(output_file)
    print(f"Borders and bold headers added and saved to {output_file}")

################################ Main ################################

input_file = 'test_data.xlsx'
output_file = 'sorted_test_data.xlsx'

# Sort data by all columns in All sheets
sort_and_save_sheets(input_file, output_file )

# # Sort data by 'ingress' and 'egress' columns in All sheets
sort_and_save_sheets(input_file, output_file , ['ingress' ,"egress"])

# # Sort Data in specific sheets using name
sort_and_save_sheets(input_file, output_file , specified_sheetname=['Sheet2'])

# Format the Excel file
add_borders_and_bold_headers_to_sheets(output_file, output_file)

# Apply autofit and centering to all columns and rows in Excel file
autofit_columns_and_rows(output_file)
