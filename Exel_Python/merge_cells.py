import pandas as pd
import openpyxl

def merge_based_on_conditions(ws, col_A, col_B):
    """
    Merges cells in column B based on contiguous matching values in both columns A and B.

    This function iterates over the specified worksheet `ws` and merges cells in `col_B`
    based on continuous ranges where both `col_A` and `col_B` have the same value.
    The merge is applied only when the values in both columns match continuously;
    any change in either column will trigger a new merge range from that point.
    The function extends its check to one row beyond the last to ensure all valid
    ranges are merged, including the final rows if applicable.

    Parameters:
    ws (openpyxl.worksheet.worksheet.Worksheet): The worksheet on which to operate.
    col_A (int): The column index of column A where the first set of values is checked.
    col_B (int): The column index of column B where the cells will be merged based on the
                 values of column A and the matching in column B itself.

    Returns:
    None: The function directly modifies the worksheet by merging cells.

    Example:
    merge_based_on_conditions(worksheet, 1, 2)  # Assuming 1 for column A and 2 for column B
    """
    # Initialize the start index to be row 1
    start_index = 1
    current_value_B = ws.cell(row=start_index, column=col_B).value
    current_value_A = ws.cell(row=start_index, column=col_A).value

    # Iterate over the rows starting from the second row 
    for row in range(2, ws.max_row + 2):  # Extend range by 1 to include the last row check
        if row <= ws.max_row:
            # Fetch values in current row for both columns
            value_B = ws.cell(row=row, column=col_B).value
            value_A = ws.cell(row=row, column=col_A).value
        else:
            value_B, value_A = None, None  # To force the last group merge

        # Check for change in value or it's beyond the last row
        if value_B != current_value_B or value_A != current_value_A:
            # Merge the previous range if valid
            if start_index < row - 1:
                ws.merge_cells(start_row=start_index, start_column=col_B, end_row=row-1, end_column=col_B)
            start_index = row
            current_value_B = value_B
            current_value_A = value_A

# Load the workbook and sheet
file_path = 'sorted_test_data.xlsx'
output_path = 'updated_test_data.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb.active  # Assumes the data is in the active sheet, modify if needed


# Specify the columns (1 for A, 2 for B, assuming A and B are the first and second columns)
merge_based_on_conditions(ws, 2,3)

# Save the workbook
wb.save(output_path)
